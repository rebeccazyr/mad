import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def load_json(file_path):
    with open(file_path, 'r') as f:
        return json.load(f)

def find_disagreements(json1, json2, json3, json4):
    # Get common keys
    common_keys = set(json1.keys()) & set(json2.keys()) & set(json3.keys()) & set(json4.keys())
    
    rows = []
    for key in sorted(common_keys, key=int):
        v1, v2, v3, v4 = json1[key], json2[key], json3[key], json4[key]
        if len(set([str(v1), str(v2), str(v3), str(v4)])) > 1:  # convert to str to avoid unhashable types
            rows.append({
                "ID": key,
                "GT": v1,
                "Retrived_Evidence": v2,
                "Full_Evidence": v3,
                "Intent_Enhanced": v4
            })
    return pd.DataFrame(rows)

def compare_and_highlight_numbers(evi_full, evi_retrieved, evi_intent_enhanced, evidence_id_to_text):
    """
    比较三个列表中的数字，返回带有颜色标记的字符串，并在每个evidence ID后添加对应的文本
    """
    # 确保输入是列表
    if not isinstance(evi_full, list):
        evi_full = []
    if not isinstance(evi_retrieved, list):
        evi_retrieved = []
    # 注意：evi_intent_enhanced 可能是字典格式，不应该强制转换为列表
    
    # 转换为集合以便快速查找
    full_set = set(evi_full)
    
    def format_with_colors_and_text(numbers, reference_set):
        if not numbers:
            return "Not Found"
        
        formatted_parts = []
        for num in numbers:
            # 获取对应的文本内容
            text_content = evidence_id_to_text.get(str(num), "Text not found")
            
            if num in reference_set:
                # 绿色标记
                formatted_parts.append(f"✓{num}: {text_content}")
            else:
                # 普通显示
                formatted_parts.append(f"✗{num}: {text_content}")
        
        return "\n".join(formatted_parts)

    def format_with_colors_intent(numbers, reference_set):
        # 检查数据类型并相应处理
        if isinstance(numbers, dict):
            # 字典格式（包含intent, pro_claim等字段）
            intent = numbers.get("intent", "N/A")
            pro_claim = numbers.get("pro_claim", "N/A")
            con_claim = numbers.get("con_claim", "N/A")
            pro_evidence_ids = numbers.get("pro_evidence_ids", [])
            con_evidence_ids = numbers.get("con_evidence_ids", [])
            
            # 格式化所有信息
            formatted_parts = []
            
            # 添加 intent 信息
            formatted_parts.append(f"Intent: {intent}")
            
            # 添加 pro_claim 信息
            formatted_parts.append(f"Pro: {pro_claim}")
            
            # 添加 con_claim 信息
            formatted_parts.append(f"Con: {con_claim}")
            
            # 格式化 pro_evidence_ids
            pro_formatted = []
            for num in pro_evidence_ids:
                text_content = evidence_id_to_text.get(str(num), "Text not found")
                
                if num in reference_set:
                    pro_formatted.append(f"✓{num}: {text_content}")
                else:
                    pro_formatted.append(f"✗{num}: {text_content}")
            if pro_formatted:
                formatted_parts.append(f"Pro_Evidence:\n" + "\n".join(pro_formatted))
            else:
                formatted_parts.append("Pro_Evidence: None")
            
            # 格式化 con_evidence_ids
            con_formatted = []
            for num in con_evidence_ids:
                text_content = evidence_id_to_text.get(str(num), "Text not found")
                
                if num in reference_set:
                    con_formatted.append(f"✓{num}: {text_content}")
                else:
                    con_formatted.append(f"✗{num}: {text_content}")
            if con_formatted:
                formatted_parts.append(f"Con_Evidence:\n" + "\n".join(con_formatted))
            else:
                formatted_parts.append("Con_Evidence: None")
            
            return "\n".join(formatted_parts)
        elif isinstance(numbers, list):
            # 列表格式，使用带文本的格式化
            return format_with_colors_and_text(numbers, reference_set)
        else:
            # 其他类型（如字符串"Not Found"）
            return str(numbers)
    
    
    return (
        format_with_colors_and_text(evi_full, full_set),
        format_with_colors_and_text(evi_retrieved, full_set),
        format_with_colors_intent(evi_intent_enhanced, full_set)
    )

if __name__ == "__main__":
    # Replace with your actual paths
    file1 = "/home/yirui/mad/chroma/test/evidence_map.json"
    file2 = "/home/yirui/mad/chroma/20_retrieved_evidence_map_bgebase.json"
    file3 = "/home/yirui/mad/intent_enhanced_retrieved_evidence.json"
    file4 = "/home/yirui/mad/chroma/test/evidence_id_to_text.json"

    json1 = load_json(file1)
    json2 = load_json(file2)
    json3 = load_json(file3)
    json4 = load_json(file4)

    with open("id_list.txt", "r") as f:
        id_list = [line.strip() for line in f if line.strip()]

    rows = []
    for id_ in id_list:
        evi_full = json1.get(id_, [])
        evi_retrieved = json2.get(id_, [])
        evi_intent_enhanced = json3.get(id_, "Not Found")
        
        # 比较数字并添加颜色标记
        full_formatted, retrieved_formatted, intent_formatted = compare_and_highlight_numbers(
            evi_full, evi_retrieved, evi_intent_enhanced, json4
        )
        
        # Split intent_formatted into text content and evidence content
        if isinstance(intent_formatted, str) and "Pro_Evidence:" in intent_formatted:
            # Split by "Pro_Evidence:" to separate text from evidence
            parts = intent_formatted.split("Pro_Evidence:")
            text_content = parts[0].strip()
            evidence_content = "Pro_Evidence:" + parts[1] if len(parts) > 1 else ""
        else:
            text_content = intent_formatted
            evidence_content = ""
        
        rows.append({
            "ID": id_, 
            "Full_Evidence_Content": full_formatted, 
            "Retrieved_Evidence_Content": retrieved_formatted, 
            "Intent_Enhanced_Text": text_content,
            "Intent_Enhanced_Evidence": evidence_content
        })

    # 创建 DataFrame
    df = pd.DataFrame(rows)
    
    # 使用 openpyxl 创建带有颜色标记的 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.title = "Selected Sentences"
    
    # 添加标题行
    headers = ["ID", "Full_Evidence_Content", "Retrieved_Evidence_Content", 
               "Intent_Enhanced_Text", "Intent_Enhanced_Evidence"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # 添加数据行
    for row_idx, row_data in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row_data["ID"])
        ws.cell(row=row_idx, column=2, value=row_data["Full_Evidence_Content"])
        ws.cell(row=row_idx, column=3, value=row_data["Retrieved_Evidence_Content"])
        ws.cell(row=row_idx, column=4, value=row_data["Intent_Enhanced_Text"])
        ws.cell(row=row_idx, column=5, value=row_data["Intent_Enhanced_Evidence"])
        
    
    # 保存带有格式化的 Excel 文件
    wb.save("selected_sentences_formatted.xlsx")
    
    # 同时保存普通的 Excel 文件
    df.to_excel("selected_sentences.xlsx", index=False)
    print("Saved to selected_sentences.xlsx and selected_sentences_formatted.xlsx")